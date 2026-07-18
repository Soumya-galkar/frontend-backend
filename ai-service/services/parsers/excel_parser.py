from openpyxl import load_workbook


def extract_text_from_excel(file_path):

    workbook = load_workbook(file_path, data_only=True)

    pages = []

    page_number = 1

    for sheet in workbook.worksheets:

        text = f"Sheet Name: {sheet.title}\n\n"

        for row in sheet.iter_rows(values_only=True):

            values = []

            for cell in row:
                if cell is None:
                    values.append("")
                else:
                    values.append(str(cell))

            text += " | ".join(values) + "\n"

        pages.append({
            "page_number": page_number,
            "text": text
        })

        page_number += 1

    return pages